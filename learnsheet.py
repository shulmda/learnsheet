#!/usr/bin/env python
# encoding=utf-8
#
# Python GUI to test memorization from an Excel Spreadsheet
#
# DJS Apr 2019
#
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import openpyxl
import random
import copy

import wx
import wx.grid

# ToDo: Remove print statements used for logging
# ToDo: Perhaps an ItemList class would be a good idea, but the methods of the standardlist class seem enough...

class Item:

    """
        Item Class for storing the Item Object
    """

    def __init__(self,row,column,value,col_heading):
        self.row = row
        self.column = column
        self.value = value
        self.col_heading = col_heading

    def GetValue(self):
        return self.value


    def GetRow(self):
        return self.row

    def GetCol(self):
        return self.column

    def getColHeading(self):
        return self.col_heading

def CheckItemIsGerman(item):
    value = item.GetValue()

    if (value.startswith("die")):
        return "die"
    if (value.startswith("der")):
        return "der"
    if (value.startswith("das")):
        return "das"
    row = item.GetRow()
    if (row.startswith("die")):
        return "die"
    if (row.startswith("der")):
        return "der"
    if (row.startswith("das")):
        return "das"
    col = item.GetCol()

    if (col.startswith("die")):
        return "die"
    if (col.startswith("der")):
        return "der"
    if (col.startswith("das")):
        return "das"

    return ""




class SampleGrid(wx.grid.Grid):
    def __init__(self, parent,pos,size):
        wx.grid.Grid.__init__(self, parent, -1)

        self.SetPosition(pos)
        self.SetSize(size)

        personalPronomen = ["ich", "du", "er/sie/es" , "wir", "ihr", "Sie/sie"]
        seinPrasens = ["bin", "bist", "ist", "sind", "seid", "sind"]
        seinPrateritum = ["war", "warst", "war", "waren", "wart", "waren"]
        self.CreateGrid(len(personalPronomen)+1, 3)
        self.SetColLabelValue(0, "A")
        self.SetColLabelValue(1, "B")
        self.SetColLabelValue(2, "C")

        self.SetRowLabelValue(0, "1")
        self.SetCellValue(0, 0, "Pronomen")
        self.SetCellValue(0, 1, "Präsens")
        self.SetCellValue(0, 2, "Präteritum")
        self.SetReadOnly(0, 0, isReadOnly=True)
        self.SetReadOnly(0, 1, isReadOnly=True)
        self.SetReadOnly(0, 2, isReadOnly=True)

        row = 0
        for i in range(0,len(personalPronomen)):
            row=row+1
            self.SetRowLabelValue(row, "%s" % (row+1))
            self.SetCellValue(row, 0, personalPronomen[i])
            self.SetCellValue(row, 1, seinPrasens[i])
            self.SetCellValue(row, 2, seinPrateritum[i])
            self.SetReadOnly(row, 0, isReadOnly=True)
            self.SetReadOnly(row, 1, isReadOnly=True)
            self.SetReadOnly(row, 2, isReadOnly=True)



class MainWindow(wx.Frame):
    def __init__(self, parent, title):
        wx.Frame.__init__(self, parent, title=title, size=(500,400), pos=(300,300))

        # Setup Window Close event
        randomId = wx.NewId()
        self.Bind(wx.EVT_MENU, self.OnCloseWindow, id=randomId)
        accel_tbl = wx.AcceleratorTable([(wx.ACCEL_CTRL, ord('W'), randomId )])
        self.SetAcceleratorTable(accel_tbl)

        # Setting up the menu.
        filemenu= wx.Menu()

        # wx.ID_ABOUT and wx.ID_EXIT are standard IDs provided by wxWidgets.
        filemenu.Append(wx.ID_EXIT,"E&xit"," Terminate the program")

        # Creating the menubar.
        menuBar = wx.MenuBar()
        self.SetMenuBar(menuBar)  # Adding the MenuBar to the Frame content.

        self.itemlist = list()

        self.lbldestination = wx.StaticText(self, label="Workbook:", pos=(20, 20))
        self.txtfilepath = wx.TextCtrl(self, value="", pos=(100, 20), size=(250,-1))

        self.btnbrowse =wx.Button(self, label="Browse", pos=(360, 18))
        self.Bind(wx.EVT_BUTTON, self.OnFindFile, self.btnbrowse)
        self.file_path = ""

        self.lblsample = wx.StaticText(self, label="Example:", pos=(20, 105))
        self.grdsample= SampleGrid(self, pos=(100, 50), size=(340, 100))



        self.lblsheetlist= wx.StaticText(self, label="Sheet List:", pos=(20, 170))

        self.sheetlist = ['<Choose Sheet>']
        self.cmbsheetlist = wx.ComboBox(self, pos=(100, 168), size=(200, -1), choices=self.sheetlist, style=wx.CB_DROPDOWN | wx.CB_READONLY)
        self.Bind(wx.EVT_COMBOBOX, self.OnSheetChange, self.cmbsheetlist)
        self.sheetname = ""
        self.max_cols = 0
        self.max_rows = 0


        self.btnlearn =wx.Button(self, label="Learn", pos=(200, 250))
        self.Bind(wx.EVT_BUTTON, self.OnLearn, self.btnlearn)
        self.btnlearn.Disable()

        self.btntest =wx.Button(self, label="Test", pos=(325, 250))
        self.Bind(wx.EVT_BUTTON, self.OnTest,self.btntest)
        self.btntest.Disable()

        self.btnshow =wx.Button(self, label="Show", pos=(70, 250))
        self.Bind(wx.EVT_BUTTON, self.OnShow,self.btnshow)
        self.btnshow.Disable()

        self.lblloadstatus = wx.StaticText(self, label="", pos=(20, 300))
        self.Centre()
        self.Show(True)

    def OnCloseWindow(self, event):
        """
            Method to close the application
        """
        self.Destroy()

    def OnLearn(self, event):
        """
            Method to create the Learning Window
        """
        LearnWindow(self, "Learn : %s" % (self.sheetname),self.itemlist)

    def OnTest(self, event):
        """
            Method to create the Testing Window
        """

        TestWindow(self, "Test : %s" % (self.sheetname),self.itemlist)


    def OnShow(self, event):
        """
            Method to create the Show Window
        """

        ShowWindow(self, "Show : %s" % (self.sheetname),self.itemlist,self.max_rows ,self.max_cols)


    def OnSheetChange(self, event):
        """
            Method to handle the changing of the sheet
        """

        self.sheetname = event.GetString()
        self.LoadWorksheet()

    def OnFindFile(self, event):
        """
            Method to open a file selection dialog
        """

        openFileDialog = wx.FileDialog(frame, "Open", "", "",
                                       "Excel files (*.xlsx)|*.xlsx",
                                       wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)

        openFileDialog.ShowModal()
        print(openFileDialog.GetPath())
        self.file_path = openFileDialog.GetPath()
        openFileDialog.Destroy()

        self.txtfilepath.AppendText(self.file_path)
        if len(self.file_path) > 0:

            print "Reading workbook...."
            self.workbook = openpyxl.load_workbook(self.file_path)
            print "Done Reading workbook...."
            self.UpdateSheetlist()


    def LoadWorksheet(self):

        """
            Method to load the worksheet selected in the list
        """

        self.itemlist = list()
        print "Finding Sheet...."
        sheet = self.workbook.get_sheet_by_name(self.sheetname)
        print "Found Sheet...."

        self.max_rows = sheet.max_row
        self.max_cols = sheet.max_column

        print "Loading List...."

        # Todo: Perhaps not memory efficient to load the row and column headings into each list entry
        #           but it is useful for when the list is randomized
        for row_num in range(2,self.max_rows+1):
            for col_num in range(2, self.max_cols + 1):
                row_heading = sheet.cell(row=row_num, column=1).value
                col_heading = sheet.cell(row=1, column=col_num).value
                first_col_heading = sheet.cell(row=1, column=1).value
                item_value = sheet.cell(row=row_num, column=col_num).value
                self.itemlist.append(Item(row_heading, col_heading, item_value,first_col_heading))

        self.lblloadstatus.SetLabel(("Loaded: %s items" % (len(self.itemlist))))
        self.btntest.Enable(enable=True)
        self.btnlearn.Enable(enable=True)
        self.btnshow.Enable(enable=True)

    def UpdateSheetlist(self):
        """
            Method to get the list of folders and populate the options Menu
        """
        self.cmbsheetlist.Clear()

        selectone = False
        for sheet in self.workbook.worksheets:
            self.cmbsheetlist.Append(sheet.title)
            if selectone == False:
                self.cmbsheetlist.Select(0)
                selectone = True
                self.sheetname = sheet.title
        self.LoadWorksheet()




class TestWindow(wx.Frame):


    def __init__(self,parent, title,itemlist):
        wx.Frame.__init__(self, wx.GetApp().TopWindow, title=title, size=(500, 400))
        randomId = wx.NewId()
        self.Bind(wx.EVT_MENU, self.OnCloseWindow, id=randomId)
        accel_tbl = wx.AcceleratorTable([(wx.ACCEL_CTRL, ord('W'), randomId )])
        self.SetAcceleratorTable(accel_tbl)




        self.safelist = copy.copy(itemlist)
        self.itemlist = copy.copy(itemlist)
        self.missedlist = list()
        random.shuffle(self.itemlist)

        self.title = title
        self.itemindex = 0
        self.misscount = 0
        self.totalmisscount = 0
        self.hitcount = 0
        self.totalhitcount = 0
        self.missedit = False
        self.review = False
        self.listcount = len(self.itemlist)



        self.lblitem = wx.StaticText(self, label="", pos=(150, 50))
        self.txtentry = wx.TextCtrl(self, value="", pos=(150, 75), size=(200, -1))


        self.btncheck = wx.Button(self, label="Check", pos=(355, 75))
        self.Bind(wx.EVT_BUTTON, self.OnCheckEntry, self.btncheck)
        self.btncheck.SetDefault()

        self.lblprogress = wx.StaticText(self, label="", pos=(150, 155))
        self.lblreview = wx.StaticText(self, label="", pos=(150, 175))
        self.lblaccuracy = wx.StaticText(self, label="", pos=(150, 195))

        self.lblhint = wx.StaticText(self, label="Hint?", pos=(150, 225))

        self.btnhint = wx.Button(self, label="Hint", pos=(50, 225))
        self.Bind(wx.EVT_BUTTON, self.OnHint, self.btnhint)



        self.btnreset = wx.Button(self, label="Reset", pos=(50, 330))
        self.Bind(wx.EVT_BUTTON, self.OnReset, self.btnreset)
        self.btnreset.Hide()

        self.btnquit = wx.Button(self, label="Quit", pos=(350, 330))
        self.Bind(wx.EVT_BUTTON, self.OnQuit, self.btnquit)
        self.btnquit.Hide()

        self.DrawIndex()
        self.Centre()
        self.Show(True)
        self.txtentry.SetFocus()

    def OnCloseWindow(self, event):
        """
            Method to Destroy the window
        """

        self.Destroy()


    def OnQuit(self, event):
        """
            Method to Close the App
        """

        self.Close()

    def OnReset(self, event):
        """
            Method to reset the status of the test and copy the original list
        """

        self.itemlist = copy.copy(self.safelist)
        self.missedlist = list()
        random.shuffle(self.itemlist)

        self.itemindex = 0
        self.misscount = 0
        self.totalmisscount = 0
        self.hitcount = 0
        self.totalhitcount = 0
        self.missedit = False
        self.review = False
        self.listcount = len(self.itemlist)
        self.btnquit.Hide()
        self.btnreset.Hide()
        self.btnhint.Enable()
        self.btncheck.Enable()

        self.DrawIndex()

        self.DrawProgress()

    def HandleMissedItem(self):

        """
            Method to handle missed responses
        """

        self.missedit = True
        self.misscount = self.misscount + 1
        if self.review == False:
            self.totalmisscount = self.totalmisscount + 1
        self.missedlist.append(self.itemlist[self.itemindex])

    def OnCheckEntry(self, event):

        """
            Method to check if the user input is correct
        """

        item = self.itemlist[self.itemindex]
        value = item.GetValue()
        print "Testing: %s" % (self.GetItemText())
        print "Correct Value: %s" % (value)

        # If the entry is correct go to the next item
        if (value == self.txtentry.GetValue()):
            print "Correct Entry: %s" % (self.txtentry.GetValue())
            if (self.missedit == False):
                self.hitcount  =self.hitcount + 1
                if self.review == False:
                    self.totalhitcount = self.totalhitcount + 1
            self.HandleNextItem()

        # If the entry is not correct, add it to the missed list
        else:
            print "Incorrect Entry: %s" % (self.txtentry.GetValue())
            if (self.missedit == False):
                self.HandleMissedItem()
            self.DrawIndex()

    def OnHint(self, event):

        """
            Method to display a hint if the user doesn't know the answer
                if the user requires a hint, add it to the missed list
        """

        item = self.itemlist[self.itemindex]
        value = item.GetValue()
        print "Correct Value: %s" % (value)

        self.lblhint.SetLabel(value)
        self.HandleMissedItem()


    def HandleNextItem(self):

        """
            Method to move on to the next item, if the end is reached move the missed list to the current list
                and start over with the missed list
        """

        self.lblhint.SetLabel("Hint?")
        self.missedit = False
        if self.itemindex + 1 < self.listcount:
            self.itemindex = self.itemindex +1
        else:
            if (len(self.missedlist) > 0):
                self.review = True
                self.itemlist = copy.copy(self.missedlist)
                self.itemindex = 0
                self.misscount = 0
                self.hitcount = 0
                self.listcount = len(self.missedlist)
                self.missedlist = list()

        self.DrawIndex()


    def DrawProgress(self):
        """
            Method to draw the progress in the labels
        """
        progresstext = "Missed: %s Correct: %s Progress: %s of %s" % (
        self.misscount, self.hitcount, self.itemindex + 1, self.listcount)

        # If not in "Review" mode, display the progress in the progress label
        if (self.review == False):
            self.lblprogress.SetLabel(progresstext)
            self.lblreview.SetLabel("")

        # If in "Review" mode, display the progress in the review label
        else:
            progresstext = "Review! %s" % (progresstext)
            self.lblreview.SetLabel(progresstext)

        # Calculate and display the accuracy of the answers
        if (self.totalmisscount + self.totalhitcount > 0):
            accuracy = (float(self.totalhitcount) / float(self.totalmisscount + self.totalhitcount)) * 100
            accuracytext = "Accuracy: %3.2f%%" % ( accuracy)
            self.lblaccuracy.SetLabel(accuracytext)
        else:
            self.lblaccuracy.SetLabel("")
    def GetItemText(self):

        """
            Method to return the item text for the current item
        """
        item = self.itemlist[self.itemindex]
        itemtext = "%s : %s" % (item.GetRow(), item.GetCol())
        return itemtext

    def DrawIndex(self):

        """
            Method to draw the current index
        """

        # If the hitcount is equal to the listcount, then there were no errors and the test is complette
        if(self.hitcount == self.listcount ):

            self.DrawProgress()
            self.btnhint.Disable()
            self.btncheck.Disable()

            self.btnquit.Show()
            self.btnreset.Show()

        else:
            self.txtentry.Clear()
            self.lblitem.SetLabel(self.GetItemText())
            self.DrawProgress()

        # Make the entry field white if the answer is correct
        if (self.missedit == False):
            self.txtentry.SetBackgroundColour('white')
        # Make the entry field red if the last answer is correct
        else:
            self.txtentry.SetBackgroundColour('red')


        # Todo: For some reason the edit box only partialy changes color Hiding and showing it, causes it to
        #   be redrawn
        self.txtentry.Hide()
        self.txtentry.Show()
        self.txtentry.SetFocus()


class ShowWindow(wx.Frame):


    def __init__(self,parent, title,itemlist,rows,cols):

        wx.Frame.__init__(self, wx.GetApp().TopWindow, title=title, size=(500, 400))
        randomId = wx.NewId()
        self.Bind(wx.EVT_MENU, self.OnCloseWindow, id=randomId)
        accel_tbl = wx.AcceleratorTable([(wx.ACCEL_CTRL, ord('W'), randomId )])
        self.SetAcceleratorTable(accel_tbl)


        self.itemlist = copy.copy(itemlist)

        self.grid = wx.grid.Grid(self)



        self.grid.CreateGrid(rows, cols)
        row = 0
        # Col headers
        item = itemlist[0]
        self.grid.SetCellValue(0, 0, item.getColHeading())
        self.grid.SetReadOnly(0, 0, isReadOnly=True)
        for i in range(0, cols-1 ):
            self.grid.SetColMinimalWidth(i+1,100)
            item = itemlist[i]
            self.grid.SetCellValue(0, i+1, item.GetCol())
            self.grid.SetReadOnly(0,i+1,isReadOnly=True)

        # Populate Cells
        row = 1
        col = 0
        for i in range(0, len(itemlist)):
            item = itemlist[i]
            if (col == 0):
                self.grid.SetCellValue(row, col, item.GetRow())
                self.grid.SetReadOnly(row, col, isReadOnly=True)
                col = 1


            self.grid.SetCellValue(row, col, item.GetValue())
            self.grid.SetReadOnly(row, col, isReadOnly=True)

            col = col + 1
            if (col == cols):
                col = 0
                row = row +1

        self.grid.AutoSizeColumns(setAsMin=False)
        width = 100
        for i in range(0, cols):
            width = width + self.grid.GetColSize(i)


        height = 70
        for i in range(0, rows):
            height = height+ self.grid.GetRowSize(i)

        height = 800 if height > 800 else height
        width = 1000 if width > 1000 else width

        self.SetSize((width,height))
        self.Centre()
        self.Show()

    def OnCloseWindow(self, event):

        """
            Method to Destroy the window
        """

        self.Destroy()


class LearnWindow(wx.Frame):


    def __init__(self,parent, title,itemlist):
        wx.Frame.__init__(self, wx.GetApp().TopWindow, title=title, size=(500, 400))
        randomId = wx.NewId()
        self.defaultcolour = self.GetBackgroundColour()
        self.Centre()



        self.itemlist = copy.copy(itemlist)
        random.shuffle(self.itemlist)

        self.title = title
        self.itemindex = 0
        self.listcount = len(self.itemlist)


        font = wx.Font(18, wx.DEFAULT, wx.NORMAL, wx.NORMAL)

        self.lblitem = wx.StaticText(self, label="", pos=(100, 50))
        self.lblitem.SetFont(font)
        self.labelvalue = wx.StaticText(self, label="", pos=(100, 80))
        self.labelvalue.SetFont(font)

        self.btnnext =wx.Button(self, label="Next", pos=(350, 110))
        self.Bind(wx.EVT_BUTTON, self.OnNextItem, self.btnnext)
        self.btnnext.SetDefault()

        self.btnshow =wx.Button(self, label="Show", pos=(200, 110))
        self.Bind(wx.EVT_BUTTON, self.OnShowItem, self.btnshow)
        self.btnshow.Hide()


        self.btnprevious =wx.Button(self, label="Previous", pos=(50, 110))
        self.Bind(wx.EVT_BUTTON, self.OnPreviousItem, self.btnprevious)
        self.lblprogress = wx.StaticText(self, label="", pos=(150, 155))

        self.chkhide = wx.CheckBox(self, label="Hide answer?", pos=(150, 200))
        self.Bind(wx.EVT_CHECKBOX, self.OnHideValue, self.chkhide)

        self.isGerman = False
        for item in self.itemlist:
            if CheckItemIsGerman(item) != "":
                self.isGerman = True
                break

        self.chkcolorcode = wx.CheckBox(self, label="Color Code Window (for German articles)?", pos=(150, 230))
        self.Bind(wx.EVT_CHECKBOX, self.OnColorCode, self.chkcolorcode)

        if (self.isGerman == True):
            self.chkcolorcode.Show(show=True)
            self.chkcolorcode.SetValue(True)
        else:
            self.chkcolorcode.Hide()

        self.Bind(wx.EVT_MENU, self.OnCloseWindow, id=randomId)
        accel_tbl = wx.AcceleratorTable([(wx.ACCEL_CTRL, ord('W'), randomId ),(wx.ACCEL_NORMAL, wx.WXK_LEFT, self.btnprevious.GetId()),(wx.ACCEL_NORMAL, wx.WXK_RIGHT, self.btnnext.GetId()),(wx.ACCEL_NORMAL, wx.WXK_DOWN, self.btnshow.GetId()) ])
        self.SetAcceleratorTable(accel_tbl)

        self.DrawIndex()
        self.Show()




    def OnCloseWindow(self, event):

        """
            Method to Destroy the window
        """

        self.Destroy()

    def OnShowItem(self, event):
        """
            Method to Show the Item Value
        """

        if (self.chkhide.IsChecked()):
            self.labelvalue.SetLabel(self.GetItemValue())


    def OnHideValue(self, event):

        """
            Method to hide the Value Label if the status changes.  The DrawIndex function checks the checbox control
        """

        self.DrawIndex()

    def OnColorCode(self, event):

        """
            Method to hide the Value Label if the status changes.  The DrawIndex function checks the checbox control
        """

        self.DrawIndex()


    def OnNextItem(self, event):

        """
            Method to move on to the next item
        """

        if self.itemindex + 1 < self.listcount:
            self.itemindex = self.itemindex +1
        self.DrawIndex()

    def OnPreviousItem(self, event):

        """
            Method to move to the previous item
        """

        if self.itemindex - 1 >= 0:
            self.itemindex = self.itemindex - 1
        self.DrawIndex()


    def DrawProgress(self):
        """
            Method to draw the progress in the label
        """

        progresstext = "Progress: %s of %s" % (self.itemindex + 1, self.listcount)

        self.lblprogress.SetLabel(progresstext)

    def GetItemText(self):

        """
            Method to return the item text for the current item
        """
        item = self.itemlist[self.itemindex]
        itemtext = "%s : %s" % (item.getColHeading(), item.GetRow())
        return itemtext

    def GetItemValue(self):

        """
            Method to return the item text for the current item
        """
        item = self.itemlist[self.itemindex]
        valuetext = "%s : %s" % (item.GetCol(), item.GetValue())
        return valuetext

    def GetItemColor(self):

        """
            Method to return the item text for the current item
        """
        itemarticle = CheckItemIsGerman(self.itemlist[self.itemindex])
        if itemarticle == "der":
            return wx.Colour(135,206,250)

        if itemarticle == "die":
            return wx.Colour(255,182,193)
        if itemarticle == "das":
            return wx.Colour(144,238,144)

        return self.defaultcolour

    def DrawIndex(self):

        """
            Method to draw the current index
        """

        self.lblitem.SetLabel(self.GetItemText())

        # if the checkbox is checked, hid the itemvalue
        if (self.chkhide.IsChecked()):
            self.labelvalue.SetLabel("")
            self.btnshow.Show(show=True)
        else:
            self.labelvalue.SetLabel(self.GetItemValue())
            self.btnshow.Hide()

        if (self.isGerman == True):
            if (self.chkcolorcode.IsChecked() == True):
                self.SetBackgroundColour(self.GetItemColor())
            else:
                self.SetBackgroundColour(self.defaultcolour)
        self.DrawProgress()





app = wx.App(False)
frame = MainWindow(None, "Learnsheet")
app.MainLoop()


